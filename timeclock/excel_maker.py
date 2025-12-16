# timeclock/excel_maker.py
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
import shutil
import re
import os


def create_default_template(save_path):
    """
    템플릿 파일이 없을 경우, 코드로 직접 예쁜 양식을 생성하는 함수
    """
    print(f"[알림] 템플릿 파일이 없어 새로 생성합니다: {save_path}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "급여명세서"

    # [페이지 설정] A4 한 장 맞춤
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.3, footer=0.3)

    # [스타일 정의]
    COLOR_HEADER = "F1F3F5"
    COLOR_TOTAL = "E8F5FF"

    line_thin = Side(style='thin', color='BDBDBD')
    line_bold = Side(style='medium', color='424242')

    border_box = Border(left=line_bold, right=line_bold, top=line_bold, bottom=line_bold)
    border_inner = Border(left=line_thin, right=line_thin, top=line_thin, bottom=line_thin)

    font_title = Font(name='맑은 고딕', size=22, bold=True, color='212529')
    font_sub = Font(name='맑은 고딕', size=14, bold=True, color='495057')
    font_head = Font(name='맑은 고딕', size=11, bold=True, color='495057')
    font_text = Font(name='맑은 고딕', size=10, color='212529')
    font_bold = Font(name='맑은 고딕', size=10, bold=True, color='212529')
    font_total = Font(name='맑은 고딕', size=12, bold=True, color='0D47A1')

    align_c = Alignment(horizontal='center', vertical='center')
    align_r = Alignment(horizontal='right', vertical='center')
    align_l_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

    fill_head = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    fill_total = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')

    # [데이터 배치]
    data = [
        (1, 1, "급 여 명 세 서", "TITLE"),
        (2, 1, "{{company}}", "SUB_TITLE"),
        (4, 1, "지급일", "LABEL"), (4, 2, "{{pay_date}}", "DATA_C"),
        (4, 4, "성 명", "LABEL"), (4, 5, "{{name}}", "DATA_C"),
        (5, 1, "지급기간", "LABEL"), (5, 2, "{{period}}", "DATA_C"),
        (5, 4, "직 급", "LABEL"), (5, 5, "사원", "DATA_C"),
        (7, 1, "지급 항목", "TH"), (7, 3, "금액", "TH"),
        (7, 4, "공제 항목", "TH"), (7, 6, "금액", "TH"),
        (8, 1, "기본급", "TD"), (8, 3, "{{base_pay}}", "TD_MONEY"),
        (8, 4, "국민연금", "TD"), (8, 6, "{{pension}}", "TD_MONEY"),
        (9, 1, "주휴수당", "TD"), (9, 3, "{{ju_hyu_pay}}", "TD_MONEY"),
        (9, 4, "건강보험", "TD"), (9, 6, "{{health_ins}}", "TD_MONEY"),
        (10, 1, "연장수당", "TD"), (10, 3, "{{overtime_pay}}", "TD_MONEY"),
        (10, 4, "장기요양", "TD"), (10, 6, "{{care_ins}}", "TD_MONEY"),
        (11, 1, "야간수당", "TD"), (11, 3, "{{night_pay}}", "TD_MONEY"),
        (11, 4, "고용보험", "TD"), (11, 6, "{{ei_ins}}", "TD_MONEY"),
        (12, 1, "휴일수당", "TD"), (12, 3, "{{holiday_pay}}", "TD_MONEY"),
        (12, 4, "소득세", "TD"), (12, 6, "{{income_tax}}", "TD_MONEY"),
        (13, 1, "기타수당", "TD"), (13, 3, "{{other_pay}}", "TD_MONEY"),
        (13, 4, "지방세", "TD"), (13, 6, "{{local_tax}}", "TD_MONEY"),
        (15, 1, "지급 합계", "TOT_L"), (15, 3, "{{total_pay}}", "TOT_M"),
        (15, 4, "공제 합계", "TOT_L"), (15, 6, "{{total_deduction}}", "TOT_M"),
        (17, 1, "실수령액 (차인지급액)", "REAL_TOT_L"), (17, 3, "{{net_pay}}", "REAL_TOT_M"),
        (19, 1, "산출 근거", "SEC_TITLE"),
        (20, 1, "{{calc_detail}}\n{{base_detail}}\n{{over_detail}}\n{{ju_hyu_detail}}", "BOX"),
        (23, 1, "비고", "SEC_TITLE"),
        (24, 1, "{{note}}", "BOX"),
        (28, 1, "위와 같이 급여를 지급합니다.", "FOOT_MSG"),
        (30, 1, "대 표 자    이  수  진   (인)", "FOOT_NAME"),
    ]

    # [렌더링]
    for r, c, val, sty in data:
        cell = ws.cell(row=r, column=c)
        cell.value = val

        if sty == "TITLE":
            cell.font = font_title
            cell.alignment = align_c
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 40
        elif sty == "SUB_TITLE":
            cell.font = font_sub
            cell.alignment = align_c
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 30
        elif sty == "LABEL":
            cell.font = font_head
            cell.fill = fill_head
            cell.alignment = align_c
            cell.border = border_inner
            ws.row_dimensions[r].height = 25
        elif sty == "DATA_C":
            cell.font = font_text
            cell.alignment = align_c
            cell.border = border_inner
            if c in [2, 5]: ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        elif sty == "TH":
            cell.font = font_head
            cell.fill = fill_head
            cell.alignment = align_c
            cell.border = border_box
            ws.row_dimensions[r].height = 30
            if c in [1, 4]: ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        elif sty == "TD":
            cell.font = font_text
            cell.alignment = align_c
            cell.border = border_inner
            ws.row_dimensions[r].height = 24
            if c in [1, 4]: ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        elif sty == "TD_MONEY":
            cell.font = font_text
            cell.alignment = align_r
            cell.border = border_inner
            cell.number_format = '#,##0'
        elif sty == "TOT_L":
            cell.font = font_bold
            cell.fill = fill_head
            cell.alignment = align_c
            cell.border = border_box
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
            ws.row_dimensions[r].height = 30
        elif sty == "TOT_M":
            cell.font = font_bold
            cell.alignment = align_r
            cell.border = border_box
            cell.number_format = '#,##0'
        elif sty == "REAL_TOT_L":
            cell.font = font_total
            cell.fill = fill_total
            cell.alignment = align_c
            cell.border = border_box
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.row_dimensions[r].height = 40
        elif sty == "REAL_TOT_M":
            cell.font = font_total
            cell.fill = fill_total
            cell.alignment = align_r
            cell.border = border_box
            cell.number_format = '#,##0'
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        elif sty == "SEC_TITLE":
            cell.font = font_bold
            cell.alignment = Alignment(horizontal='left', vertical='bottom')
            cell.border = Border(bottom=line_bold)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 30
        elif sty == "BOX":
            cell.font = font_text
            cell.alignment = align_l_top
            ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
            for rr in range(r, r + 3):
                for cc in range(1, 7):
                    ws.cell(rr, cc).border = border_inner
            ws.row_dimensions[r].height = 70
        elif sty == "FOOT_MSG":
            cell.font = font_text
            cell.alignment = align_c
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 30
        elif sty == "FOOT_NAME":
            cell.font = Font(name='맑은 고딕', size=16, bold=True)
            cell.alignment = align_c
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 50

    # [열 너비 설정]
    widths = {'A': 13, 'B': 13, 'C': 22, 'D': 13, 'E': 13, 'F': 22}
    for col_char, width in widths.items():
        ws.column_dimensions[col_char].width = width

    wb.save(save_path)


def generate_payslip(template_path, save_path, data_context):
    """
    1. 템플릿이 없으면 자동 생성
    2. 데이터 치환 및 저장
    """
    # ★ 템플릿 파일 자동 복구 기능
    if not os.path.exists(template_path):
        try:
            create_default_template(template_path)
        except Exception as e:
            print(f"[오류] 템플릿 생성 실패: {e}")
            return None

    print(f"\n[엑셀 생성 시작] {save_path}")

    # 1. 템플릿 복사
    try:
        shutil.copy(template_path, save_path)
    except Exception as e:
        print(f"[오류] 파일 복사 실패: {e}")
        return None

    # 2. 엑셀 로드
    try:
        wb = openpyxl.load_workbook(save_path)
        ws = wb.active
    except Exception as e:
        print(f"[오류] 엑셀 로드 실패: {e}")
        return None

    # 3. 치환 작업
    replaced_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                text = str(cell.value)
                for key, val in data_context.items():
                    pattern = r"\{\{\s*" + re.escape(key) + r"\s*\}\}"
                    if re.search(pattern, text):
                        if re.fullmatch(pattern, text.strip()):
                            cell.value = val
                        else:
                            cell.value = re.sub(pattern, str(val), text)

                        replaced_count += 1
                        text = str(cell.value)

    # 4. 저장
    try:
        wb.save(save_path)
        wb.close()
        print(f"[완료] 총 {replaced_count}개 항목 입력됨.")
    except Exception as e:
        print(f"[오류] 저장 실패: {e}")
        return None

    return str(save_path)