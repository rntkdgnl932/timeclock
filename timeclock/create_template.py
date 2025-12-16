# create_template.py
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
import os

# 1. 저장 경로
base_dir = r"C:\my_games\timeclock\app_data"
if not os.path.exists(base_dir):
    os.makedirs(base_dir)

save_path = os.path.join(base_dir, "template.xlsx")

# 2. 엑셀 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "급여명세서"

# ==========================================
# [★핵심] 중앙 정렬 및 페이지 설정
# ==========================================
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

# 인쇄 시 가로/세로 가운데 정렬 (이게 왼쪽 쏠림 해결의 핵심!)
ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = False  # 세로는 위쪽부터 채움

# 한 페이지에 맞춤
ws.page_setup.fitToPage = True
ws.page_setup.fitToHeight = 1
ws.page_setup.fitToWidth = 1

# 여백 조정 (양쪽 여백을 균일하게 줌)
ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.3, footer=0.3)

# ==========================================
# [스타일] 모던 & 심플
# ==========================================
# 색상
COLOR_HEADER = "F1F3F5"  # 연한 회색
COLOR_TOTAL = "E8F5FF"  # 연한 블루

# 테두리
line_thin = Side(style='thin', color='BDBDBD')
line_bold = Side(style='medium', color='424242')

border_box = Border(left=line_bold, right=line_bold, top=line_bold, bottom=line_bold)
border_inner = Border(left=line_thin, right=line_thin, top=line_thin, bottom=line_thin)
border_bottom = Border(bottom=line_bold)

# 폰트
font_title = Font(name='맑은 고딕', size=22, bold=True, color='212529')
font_sub = Font(name='맑은 고딕', size=14, bold=True, color='495057')
font_head = Font(name='맑은 고딕', size=11, bold=True, color='495057')
font_text = Font(name='맑은 고딕', size=10, color='212529')
font_bold = Font(name='맑은 고딕', size=10, bold=True, color='212529')
font_total = Font(name='맑은 고딕', size=12, bold=True, color='0D47A1')

# 정렬
align_c = Alignment(horizontal='center', vertical='center')
align_r = Alignment(horizontal='right', vertical='center')
align_l_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

# 채우기
fill_head = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
fill_total = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')

# ==========================================
# [데이터 배치]
# ==========================================
data = [
    # 행, 열, 값, 스타일
    # 1. 제목 및 회사명
    (1, 1, "급 여 명 세 서", "TITLE"),
    (2, 1, "{{company}}", "SUB_TITLE_CENTER"),

    # 2. 기본 정보
    (4, 1, "지급일", "LABEL"), (4, 2, "{{pay_date}}", "DATA_C"),
    (4, 4, "성 명", "LABEL"), (4, 5, "{{name}}", "DATA_C"),
    (5, 1, "지급기간", "LABEL"), (5, 2, "{{period}}", "DATA_C"),
    (5, 4, "직 급", "LABEL"), (5, 5, "사원", "DATA_C"),

    # 3. 헤더
    (7, 1, "지급 항목", "TH"), (7, 3, "금액", "TH"),
    (7, 4, "공제 항목", "TH"), (7, 6, "금액", "TH"),

    # 4. 상세 내역
    (8, 1, "기본급", "TD"), (8, 3, "{{base_pay}}", "TD_MONEY"),
    (8, 4, "국민연금", "TD"), (8, 6, "{{pension}}", "TD_MONEY"),

    (9, 1, "주휴수당", "TD"), (9, 3, "{{ju_hyu_pay}}", "TD_MONEY"),
    (9, 4, "건강보험", "TD"), (9, 6, "{{health_ins}}", "TD_MONEY"),

    (10, 1, "연장수당", "TD"), (10, 3, "{{overtime_pay}}", "TD_MONEY"),
    (10, 4, "장기요양", "TD"), (10, 6, "{{care_ins}}", "TD_MONEY"),

    (11, 1, "야간수당", "TD"), (11, 3, "{{night_pay}}", "TD_MONEY"),
    (11, 4, "고용보험", "TD"), (11, 6, "{{ei_ins}}", "TD_MONEY"),

    (12, 1, "휴일수당", "TD"), (12, 3, "{{holiday_pay}}", "TD_MONEY"),
    (12, 4, "소득세", "TD"), (12, 6, "{{income_tax}}", "TD_MONEY"),

    (13, 1, "기타수당", "TD"), (13, 3, "{{other_pay}}", "TD_MONEY"),
    (13, 4, "지방세", "TD"), (13, 6, "{{local_tax}}", "TD_MONEY"),

    # 5. 합계
    (15, 1, "지급 합계", "TOT_L"), (15, 3, "{{total_pay}}", "TOT_M"),
    (15, 4, "공제 합계", "TOT_L"), (15, 6, "{{total_deduction}}", "TOT_M"),

    # 6. 실수령액
    (17, 1, "실수령액 (차인지급액)", "REAL_TOT_L"), (17, 3, "{{net_pay}}", "REAL_TOT_M"),

    # 7. 산출근거 & 비고
    (19, 1, "산출 근거", "SEC_TITLE"),
    (20, 1, "{{calc_detail}}\n{{base_detail}}\n{{over_detail}}\n{{ju_hyu_detail}}", "BOX"),

    (23, 1, "비고", "SEC_TITLE"),
    (24, 1, "{{note}}", "BOX"),

    # 8. 하단 서명
    (28, 1, "위와 같이 급여를 지급합니다.", "FOOT_MSG"),
    (30, 1, "대 표 자    이  수  진   (인)", "FOOT_NAME"),
]

# ==========================================
# [렌더링 루프]
# ==========================================
for item in data:
    r, c, val, sty = item
    cell = ws.cell(row=r, column=c)
    cell.value = val

    if sty == "TITLE":
        cell.font = font_title
        cell.alignment = align_c
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 40

    elif sty == "SUB_TITLE_CENTER":
        cell.font = font_sub
        cell.alignment = align_c
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 30

    elif sty == "LABEL":
        cell.font = font_head
        cell.fill = fill_head
        cell.alignment = align_c
        cell.border = border_inner
        ws.row_dimensions[r].height = 25

    elif sty == "DATA_C":
        cell.font = font_text
        cell.alignment = align_c
        cell.border = border_inner
        if c == 2 or c == 5:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)

    elif sty == "TH":
        cell.font = font_head
        cell.fill = fill_head
        cell.alignment = align_c
        cell.border = border_box
        ws.row_dimensions[r].height = 30
        if c == 1 or c == 4:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)

    elif sty == "TD":
        cell.font = font_text
        cell.alignment = align_c
        cell.border = border_inner
        ws.row_dimensions[r].height = 24
        if c == 1 or c == 4:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)

    elif sty == "TD_MONEY":
        cell.font = font_text
        cell.alignment = align_r
        cell.border = border_inner
        cell.number_format = '#,##0'

    elif sty == "TOT_L":
        cell.font = font_bold
        cell.fill = fill_head
        cell.alignment = align_c
        cell.border = border_box
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        ws.row_dimensions[r].height = 30

    elif sty == "TOT_M":
        cell.font = font_bold
        cell.alignment = align_r
        cell.border = border_box
        cell.number_format = '#,##0'

    elif sty == "REAL_TOT_L":
        cell.font = font_total
        cell.fill = fill_total
        cell.alignment = align_c
        cell.border = border_box
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.row_dimensions[r].height = 40

    elif sty == "REAL_TOT_M":
        cell.font = font_total
        cell.fill = fill_total
        cell.alignment = align_r
        cell.border = border_box
        cell.number_format = '#,##0'
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)

    elif sty == "SEC_TITLE":
        cell.font = font_bold
        cell.alignment = Alignment(horizontal='left', vertical='bottom')
        cell.border = Border(bottom=line_bold)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 30

    elif sty == "BOX":
        cell.font = font_text
        cell.alignment = align_l_top
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
        for rr in range(r, r + 3):
            for cc in range(1, 7):
                ws.cell(rr, cc).border = border_inner
        ws.row_dimensions[r].height = 70

    elif sty == "FOOT_MSG":
        cell.font = font_text
        cell.alignment = align_c
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 30

    elif sty == "FOOT_NAME":
        cell.font = Font(name='맑은 고딕', size=16, bold=True)
        cell.alignment = align_c
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 50

# ==========================================
# [★핵심] 열 너비 확장 (페이지 꽉 채우기)
# ==========================================
# 기존보다 너비를 확 늘려서 A4 가로폭을 꽉 채우게 만듭니다.
# A,D(항목명): 13
# B,E(데이터): 13
# C,F(금액): 22 (금액란을 특히 넓게)
ws.column_dimensions['A'].width = 13
ws.column_dimensions['B'].width = 13
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 13
ws.column_dimensions['E'].width = 13
ws.column_dimensions['F'].width = 22

wb.save(save_path)
print(f"✨ [완료] 중앙 정렬 & 너비 꽉 채움 완료!\n   경로: {save_path}")